"""Excel import helpers for admin bulk data (locations, inspectors).

Templates use a header row plus an optional example/notes row; data rows follow.
Column headers accept zh-TW labels or English field names.
"""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy.orm import Session

from .config import get_settings
from .models import Inspector, Location
from .security import hash_password

logger = logging.getLogger("parking.import")

IMPORT_TYPES = frozenset({"locations", "inspectors"})

FIELD_LABELS = {
    "district": "行政區",
    "road": "路段",
    "spot_no": "停車格編號",
    "username": "帳號",
    "password": "密碼",
    "display_name": "姓名",
}

# Column widths from models.py. SQLite ignores VARCHAR lengths, so an over-long
# cell only fails on PostgreSQL — and it fails as a DataError, which is *not* an
# IntegrityError: uncaught, it aborts the import with a 500, leaving the rows
# committed before it applied and the admin with no report of what landed.
# Checking here turns that into an ordinary per-row error. VARCHAR(n) counts
# characters, not bytes, so len() is the right measure.
LOCATION_MAX_LENGTHS = {"district": 64, "road": 128, "spot_no": 64}
INSPECTOR_MAX_LENGTHS = {"username": 64, "display_name": 128}

LOCATION_HEADER_MAP = {
    "行政區": "district",
    "district": "district",
    "路段": "road",
    "road": "road",
    "停車格編號": "spot_no",
    "停車格": "spot_no",
    "spot_no": "spot_no",
}

INSPECTOR_HEADER_MAP = {
    "帳號": "username",
    "username": "username",
    "密碼": "password",
    "password": "password",
    "姓名": "display_name",
    "display_name": "display_name",
    "啟用權限": "has_permission",
    "has_permission": "has_permission",
}

LOCATION_REQUIRED = frozenset({"district", "road", "spot_no"})
INSPECTOR_REQUIRED = frozenset({"username", "password", "display_name"})


# A file where every row is wrong would otherwise return one JSON entry per row
# and have the console render a table of all of them. Report enough to fix the
# file, count the rest.
MAX_REPORTED_ERRORS = 200


@dataclass
class ImportResult:
    import_type: str
    total_rows: int
    created: int
    skipped: int
    error_count: int = 0
    errors: list[dict[str, Any]] = field(default_factory=list)

    def add_error(self, row: int, message: str) -> None:
        self.error_count += 1
        if len(self.errors) < MAX_REPORTED_ERRORS:
            self.errors.append({"row": row, "message": message})

    def to_dict(self) -> dict[str, Any]:
        return {
            "import_type": self.import_type,
            "total_rows": self.total_rows,
            "created": self.created,
            "skipped": self.skipped,
            "error_count": self.error_count,
            "errors": self.errors,
            "errors_truncated": self.error_count > len(self.errors),
        }


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_example_row(cells: tuple[Any, ...]) -> bool:
    first = _cell_text(cells[0]) if cells else ""
    return first.startswith("範例") or first.startswith("示例") or first.lower().startswith("example")


def _row_is_empty(cells: tuple[Any, ...]) -> bool:
    return all(not _cell_text(c) for c in cells)


def _length_error(values: dict[str, str], caps: dict[str, int]) -> str | None:
    for key, cap in caps.items():
        if len(values.get(key, "")) > cap:
            return f"{FIELD_LABELS[key]}長度超過 {cap} 字元"
    return None


def _parse_permission(value: Any) -> tuple[bool | None, str | None]:
    text = _cell_text(value)
    if not text:
        return True, None
    lowered = text.lower()
    if lowered in {"是", "1", "true", "yes", "y", "啟用", "有"}:
        return True, None
    if lowered in {"否", "0", "false", "no", "n", "停用", "無"}:
        return False, None
    return None, f"啟用權限格式無效：{text}（請填 是/否）"


def parse_workbook_rows(content: bytes, import_type: str) -> tuple[list[tuple[int, dict[str, str]]], str | None]:
    """Return (rows with 1-based Excel row numbers, fatal parse error)."""
    header_map = LOCATION_HEADER_MAP if import_type == "locations" else INSPECTOR_HEADER_MAP
    required = LOCATION_REQUIRED if import_type == "locations" else INSPECTOR_REQUIRED
    row_limit = get_settings().max_import_rows

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return [], "無法讀取 Excel 檔案，請確認格式為 .xlsx"

    try:
        # `active` is None for a workbook whose activeTab points past the last
        # sheet — a real .xlsx openpyxl loads happily, so treat it as an
        # unreadable file rather than letting it AttributeError into a 500.
        ws = wb.active
        if ws is None:
            return [], "無法讀取 Excel 檔案，請確認格式為 .xlsx"

        # Stream the sheet instead of list()ing it: a 5 MB workbook can hold
        # ~500k rows, and materialising those costs ~215 MB before a single
        # value is validated. Scanning stops at `row_limit` for the same
        # reason.
        col_keys: list[str | None] = []
        header_found = False
        saw_any_row = False
        over_limit = False
        parsed: list[tuple[int, dict[str, str]]] = []

        for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_num > row_limit:
                over_limit = True
                break
            saw_any_row = True
            if _row_is_empty(row) or _is_example_row(row):
                continue

            if not header_found:
                labels = [_normalize_header(c) for c in row]
                mapped = [header_map.get(label) for label in labels]
                if any(k in required for k in mapped if k):
                    header_found = True
                    col_keys = mapped
                continue

            record: dict[str, str] = {}
            for key, cell in zip(col_keys, row):
                if key:
                    record[key] = _cell_text(cell)
            if any(record.get(k) for k in required):
                parsed.append((row_num, record))
    finally:
        # A read_only workbook holds the underlying zip open until closed.
        wb.close()

    if over_limit:
        return [], f"檔案列數超過上限 {row_limit} 列（含標題與空白列），請分批匯入"
    if not saw_any_row:
        return [], "檔案沒有資料列"
    if not header_found:
        return [], "找不到有效的欄位標題列（請使用系統提供的範本）"

    missing = sorted(required - {k for k in col_keys if k})
    if missing:
        return [], f"缺少必要欄位：{', '.join(FIELD_LABELS[k] for k in missing)}"

    if not parsed:
        return [], "沒有可匯入的資料列"

    return parsed, None


def build_template_workbook(import_type: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    if import_type == "locations":
        ws.title = "停車格匯入"
        ws.append(["行政區", "路段", "停車格編號"])
        ws.append(["範例：信義區", "松高路", "Z-001"])
    else:
        ws.title = "稽查員匯入"
        ws.append(["帳號", "密碼", "姓名", "啟用權限"])
        ws.append(["範例：insp99", "pass123", "測試員", "是"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _commit_row(db: Session, result: ImportResult, row_num: int) -> bool:
    """Commit one pending row. Returns True if it landed.

    A duplicate is an expected outcome (skip). Anything else is a row we can't
    write — record it and move on rather than letting it escape as a 500 that
    leaves the file half-applied with no report.
    """
    try:
        db.commit()
        return True
    except IntegrityError:
        db.rollback()
        result.skipped += 1
        return False
    except SQLAlchemyError:
        db.rollback()
        logger.exception("import: row %s failed to write", row_num)
        result.add_error(row_num, "寫入資料庫失敗，已略過此列")
        return False


def import_locations(db: Session, rows: list[tuple[int, dict[str, str]]]) -> ImportResult:
    result = ImportResult(import_type="locations", total_rows=len(rows), created=0, skipped=0)
    for row_num, row in rows:
        district = row.get("district", "").strip()
        road = row.get("road", "").strip()
        spot_no = row.get("spot_no", "").strip()
        if not district or not road or not spot_no:
            result.add_error(row_num, "行政區、路段、停車格編號皆為必填")
            continue

        too_long = _length_error(
            {"district": district, "road": road, "spot_no": spot_no}, LOCATION_MAX_LENGTHS
        )
        if too_long:
            result.add_error(row_num, too_long)
            continue

        existing = db.scalar(
            select(Location).where(
                Location.district == district,
                Location.road == road,
                Location.spot_no == spot_no,
            )
        )
        if existing:
            result.skipped += 1
            continue

        db.add(Location(district=district, road=road, spot_no=spot_no))
        if _commit_row(db, result, row_num):
            result.created += 1
    return result


def import_inspectors(db: Session, rows: list[tuple[int, dict[str, str]]]) -> ImportResult:
    result = ImportResult(import_type="inspectors", total_rows=len(rows), created=0, skipped=0)
    for row_num, row in rows:
        username = row.get("username", "").strip()
        password = row.get("password", "").strip()
        display_name = row.get("display_name", "").strip()
        if not username or not password or not display_name:
            result.add_error(row_num, "帳號、密碼、姓名皆為必填")
            continue

        too_long = _length_error(
            {"username": username, "display_name": display_name}, INSPECTOR_MAX_LENGTHS
        )
        if too_long:
            result.add_error(row_num, too_long)
            continue

        has_permission, perm_err = _parse_permission(row.get("has_permission"))
        if perm_err:
            result.add_error(row_num, perm_err)
            continue

        existing = db.scalar(select(Inspector).where(Inspector.username == username))
        if existing:
            result.skipped += 1
            continue

        db.add(
            Inspector(
                username=username,
                password=hash_password(password),
                display_name=display_name,
                has_permission=int(has_permission),
            )
        )
        if _commit_row(db, result, row_num):
            result.created += 1
    return result


def run_import(db: Session, import_type: str, content: bytes) -> tuple[ImportResult | None, str | None]:
    rows, parse_error = parse_workbook_rows(content, import_type)
    if parse_error:
        return None, parse_error
    if import_type == "locations":
        return import_locations(db, rows), None
    return import_inspectors(db, rows), None
