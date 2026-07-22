"""CSV and Excel export match the field team's spreadsheet column layout."""
from __future__ import annotations

import csv
import io

from openpyxl import load_workbook

from tests.conftest import auth


def _make_case(client, inspector_token, **overrides):
    payload = {
        "ticket_no": "Q7036002A121045",  # -> Q703 | 6002A | 1210 | 45
        "district": "大安區",
        "road": "敦化南路",
        "spot_no": "C-101",
        "plate_no": "GHI-3456",
        "amount": 900,
        "due_date": "2026-07-24",
        "parking_date": "2026-07-03",
        "parking_start": "2026-07-03T11:40:00",
        "parking_end": "2026-07-03T12:40:00",
        "data_source": "AUTO_QR",
        "inspector_username": "insp01",
    }
    payload.update(overrides)
    res = client.post("/api/cases", headers=auth(inspector_token), json=payload)
    assert res.status_code == 200, res.text
    return res.json()


def test_export_headers_and_barcode_split(client, inspector_token, manager_token):
    _make_case(client, inspector_token)
    res = client.get("/api/admin/export.csv", headers=auth(manager_token))
    assert res.status_code == 200

    text = res.text
    if text and text[0] == "﻿":  # strip BOM
        text = text[1:]
    rows = list(csv.reader(io.StringIO(text)))

    assert rows[0] == [
        "日期", "檢查時間", "調查員", "路段", "停車格編號", "車號",
        "日期", "調查員", "時間", "秒", "可不用", "費率", "其他",
    ]
    assert rows[1][0] == "稽查當下日期"
    assert "條碼下方數字" in rows[1][6]

    # Find the row for the case we created (by plate).
    mine = [r for r in rows[2:] if r[5] == "GHI-3456"]
    assert mine, "created case missing from export"
    d = mine[0]
    assert d[2] == "王小明"      # C 調查員 = inspector display name (insp01)
    assert d[3] == "敦化南路"    # D 路段
    assert d[4] == "C-101"       # E 停車格編號
    assert d[6] == "Q703"        # G 日期 (barcode)
    assert d[7] == "6002A"       # H 調查員 (barcode)
    assert d[8] == "1210"        # I 時間 (barcode)
    assert d[9] == "45"          # J 秒 (barcode)
    assert d[10] == "" and d[11] == "" and d[12] == ""  # K/L/M reserved


def test_export_neutralises_formula_injection(client, inspector_token, manager_token):
    # A plate crafted to run as a spreadsheet formula must be exported as
    # literal text (prefixed with ') so it can't execute in a reviewer's Excel.
    _make_case(client, inspector_token, plate_no="=1+2")
    res = client.get("/api/admin/export.csv", headers=auth(manager_token))
    assert res.status_code == 200

    text = res.text
    if text and text[0] == "﻿":  # strip BOM
        text = text[1:]
    rows = list(csv.reader(io.StringIO(text)))

    mine = [r for r in rows[2:] if r[5] == "'=1+2"]
    assert mine, "formula-injection plate was not neutralised in the export"


def test_export_xlsx_headers_and_content(client, inspector_token, manager_token):
    _make_case(client, inspector_token)
    res = client.get("/api/admin/export.xlsx", headers=auth(manager_token))
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]

    wb = load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws.cell(1, 1).value == "日期"
    assert ws.cell(2, 1).value == "稽查當下日期"

    plates = [ws.cell(row, 6).value for row in range(3, ws.max_row + 1)]
    assert "GHI-3456" in plates


def test_export_respects_filters(client, inspector_token, manager_token):
    _make_case(client, inspector_token, district="大安區", plate_no="FILTER-A")
    _make_case(
        client,
        inspector_token,
        ticket_no="Q7036002A121046",
        district="信義區",
        road="信義路",
        plate_no="FILTER-B",
    )
    res = client.get("/api/admin/export.csv?district=大安區", headers=auth(manager_token))
    assert res.status_code == 200

    text = res.text
    if text and text[0] == "﻿":
        text = text[1:]
    rows = list(csv.reader(io.StringIO(text)))
    plates = [r[5] for r in rows[2:]]
    assert "FILTER-A" in plates
    assert "FILTER-B" not in plates
