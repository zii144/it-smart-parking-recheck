"""Admin Excel import: templates, bulk locations/inspectors, RBAC."""
from __future__ import annotations

import inspect
import io
import re
import zipfile

from openpyxl import Workbook

from app.config import get_settings
from app.main import admin_import_excel
from tests.conftest import auth

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _post_import(client, token, xlsx: bytes, import_type: str = "locations", name: str = "spots.xlsx"):
    return client.post(
        f"/api/admin/import/{import_type}",
        headers=auth(token),
        files={"file": (name, xlsx, XLSX_MIME)},
    )


def test_import_endpoint_runs_off_the_event_loop():
    """The handler must stay a sync `def` so Starlette runs it in the threadpool.

    Its work is blocking (openpyxl parse, per-row DB commit, ~0.3s of bcrypt per
    inspector). As an `async def` that would run on the event loop and stall
    every other request for the whole import.
    """
    assert not inspect.iscoroutinefunction(admin_import_excel)


def test_import_template_locations(client, sysadmin_token):
    res = client.get("/api/admin/import/templates/locations", headers=auth(sysadmin_token))
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]
    assert "parking_locations_import_template.xlsx" in res.headers["content-disposition"]


def test_import_template_inspectors(client, sysadmin_token):
    res = client.get("/api/admin/import/templates/inspectors", headers=auth(sysadmin_token))
    assert res.status_code == 200
    assert "parking_inspectors_import_template.xlsx" in res.headers["content-disposition"]


def test_import_template_rejects_manager(client, manager_token):
    res = client.get("/api/admin/import/templates/locations", headers=auth(manager_token))
    assert res.status_code == 403


def test_import_unknown_type(client, sysadmin_token):
    res = client.get("/api/admin/import/templates/cases", headers=auth(sysadmin_token))
    assert res.status_code == 400


def test_import_locations_success_and_skip_duplicate(client, sysadmin_token):
    xlsx = _build_xlsx([
        ["行政區", "路段", "停車格編號"],
        ["信義區", "松高路", "IMP-001"],
        ["信義區", "松高路", "IMP-001"],  # duplicate row -> skipped
        ["大安區", "敦化南路", "IMP-002"],
    ])
    res = client.post(
        "/api/admin/import/locations",
        headers=auth(sysadmin_token),
        files={"file": ("spots.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    body = res.json()
    assert body["import_type"] == "locations"
    assert body["total_rows"] == 3
    assert body["created"] == 2
    assert body["skipped"] == 1
    assert body["errors"] == []

    listed = client.get("/api/admin/locations", headers=auth(sysadmin_token)).json()
    spots = {(r["district"], r["road"], r["spot_no"]) for r in listed}
    assert ("信義區", "松高路", "IMP-001") in spots
    assert ("大安區", "敦化南路", "IMP-002") in spots


def test_import_locations_validation_errors(client, sysadmin_token):
    xlsx = _build_xlsx([
        ["行政區", "路段", "停車格編號"],
        ["", "松高路", "BAD-001"],
        ["大安區", "敦化南路", "BAD-002"],
    ])
    res = client.post(
        "/api/admin/import/locations",
        headers=auth(sysadmin_token),
        files={"file": ("spots.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    body = res.json()
    assert body["created"] == 1
    assert body["errors"] == [{"row": 2, "message": "行政區、路段、停車格編號皆為必填"}]


def test_import_locations_rejects_non_xlsx(client, sysadmin_token):
    res = client.post(
        "/api/admin/import/locations",
        headers=auth(sysadmin_token),
        files={"file": ("spots.csv", b"a,b,c", "text/csv")},
    )
    assert res.status_code == 400


# --- resource limits --------------------------------------------------------
def test_import_rejects_oversized_upload(client, sysadmin_token, monkeypatch):
    """An .xlsx compresses hard, so bytes-on-the-wire must be capped too."""
    monkeypatch.setattr(get_settings(), "max_import_bytes", 1024)
    xlsx = _build_xlsx([["行政區", "路段", "停車格編號"]] + [["信義區", "松高路", f"BIG-{i}"] for i in range(500)])
    assert len(xlsx) > 1024

    res = _post_import(client, sysadmin_token, xlsx)
    assert res.status_code == 413
    assert "過大" in res.json()["detail"]


def test_import_rejects_too_many_rows(client, sysadmin_token, monkeypatch):
    """Scanning is bounded so a huge sheet can't tie up a worker for minutes."""
    monkeypatch.setattr(get_settings(), "max_import_rows", 5)
    xlsx = _build_xlsx([["行政區", "路段", "停車格編號"]] + [["信義區", "松高路", f"MANY-{i}"] for i in range(20)])

    res = _post_import(client, sysadmin_token, xlsx)
    assert res.status_code == 400
    assert "超過上限" in res.json()["detail"]

    # Nothing was written: the cap is enforced during parsing, before import.
    listed = client.get("/api/admin/locations", headers=auth(sysadmin_token)).json()
    assert not any(r["spot_no"].startswith("MANY-") for r in listed)


def test_import_accepts_a_file_right_at_the_row_limit(client, sysadmin_token, monkeypatch):
    monkeypatch.setattr(get_settings(), "max_import_rows", 4)
    xlsx = _build_xlsx([
        ["行政區", "路段", "停車格編號"],
        ["信義區", "松高路", "EDGE-1"],
        ["信義區", "松高路", "EDGE-2"],
        ["信義區", "松高路", "EDGE-3"],
    ])

    res = _post_import(client, sysadmin_token, xlsx)
    assert res.status_code == 200, res.text
    assert res.json()["created"] == 3


def test_import_rejects_workbook_with_no_active_sheet(client, sysadmin_token):
    """`wb.active` is None when activeTab points past the last sheet.

    openpyxl loads such a file happily, so without a guard the parser would
    AttributeError into a 500 instead of a readable 400.
    """
    wb = Workbook()
    wb.active.append(["行政區", "路段", "停車格編號"])
    wb.create_sheet("two")
    buf = io.BytesIO()
    wb.save(buf)

    src = zipfile.ZipFile(io.BytesIO(buf.getvalue()))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "xl/workbook.xml":
                data = re.sub(rb"<workbookView[^>]*/>", b'<workbookView activeTab="7"/>', data)
            dst.writestr(item, data)

    res = _post_import(client, sysadmin_token, out.getvalue())
    assert res.status_code == 400
    assert "無法讀取" in res.json()["detail"]


def test_import_inspectors_success(client, sysadmin_token):
    xlsx = _build_xlsx([
        ["帳號", "密碼", "姓名", "啟用權限"],
        ["imp_x1", "secret1", "匯入員一", "是"],
        ["imp_x2", "secret2", "匯入員二", "否"],
    ])
    res = client.post(
        "/api/admin/import/inspectors",
        headers=auth(sysadmin_token),
        files={"file": ("inspectors.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    body = res.json()
    assert body["created"] == 2
    assert body["skipped"] == 0

    listed = client.get("/api/admin/inspectors", headers=auth(sysadmin_token)).json()
    by_name = {r["username"]: r for r in listed}
    assert by_name["imp_x1"]["has_permission"] == 1
    assert by_name["imp_x2"]["has_permission"] == 0

    login = client.post("/api/login", json={"username": "imp_x1", "password": "secret1"})
    assert login.status_code == 200


def test_import_inspectors_skip_existing(client, sysadmin_token):
    xlsx = _build_xlsx([
        ["帳號", "密碼", "姓名"],
        ["insp01", "newpass", "重複帳號"],
    ])
    res = client.post(
        "/api/admin/import/inspectors",
        headers=auth(sysadmin_token),
        files={"file": ("inspectors.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    body = res.json()
    assert body["created"] == 0
    assert body["skipped"] == 1
